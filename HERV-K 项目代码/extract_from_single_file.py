"""
合并所有vcf文件：
思路：
1.遍历所有vcf文件，先把所有重复的位点信息剔除并排序
"""
import os
from openpyxl import load_workbook


# 这个函数会生成按照变异位点的不同进行区分的、详细的hervk位点列表，但是没有individual的信息
# def detailed_position_list():
#     filelist = os.listdir(r'G:\项目-HERV\individual-analysis\merge_vcf_file')
#     f1 = open('G:/项目-HERV/individual-analysis/merge_single_file.vcf', "a")
#     for file in filelist:
#         ftemp = open('G:/项目-HERV/individual-analysis/merge_vcf_file/'+file, "r")
#         for i in ftemp.readlines():
#             i = i.strip()
#             if i.startswith("##"):
#                 pass
#             else:
#                 colum = i.split("\t")
#                 if 'ac0' in colum[6]:
#                     pass
#                 else:
#                     if colum[0] == '#CHROM':
#                         header = 'position\tlenth\tALT\tstrand\tfilter\tTSD\tlocation\tvariation\tgenotype\t' + colum[-1]
#                         f1.write(header+'\n')
#                     else:
#                         position = colum[0] + "_" + colum[1] + "-" + str(int(colum[1])+int(colum[7].split(';')[4].split('=')[1]))
#                         lenth = colum[7].split(';')[4].split('=')[1]
#                         ALT = colum[3]
#                         strand = colum[7].split(";")[5].split(',')[3]
#                         filter = colum[6]
#                         TSD = colum[7].split(";")[0].split('=')[1]
#                         location = colum[7].split(";")[5].split(',')[1] + '--' + colum[7].split(";")[5].split(',')[2]
#                         variation = colum[7].split(";")[6]
#                         genotype = colum[9]
#                         f1.write(position+'\t'+lenth+"\t"+ALT+"\t"+strand+"\t"+filter+"\t"+TSD+"\t"+location+"\t"+variation+'\t'+genotype+'\n')
#         ftemp.close()
#     f1.close()




# 查找特定字符串在excel表中的位置，返回值有三个，分别为 字符串、第几行、第几列
def search_value(keyword):
    wb = load_workbook("G:/项目-HERV/individual-analysis/position_list2.xlsx")
    all_sheets = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(all_sheets[0])
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                info = str(cell.value).find(keyword)
                if info == 0:
                    return cell.value, cell.row, cell.column

# 生成一个简单的只包含位置信息，不包含变异位点信息的hervklist，同时返回一个二维字典
def simple_position_list():
    hervklist = []  # 存储每一条hervk位点信息，用来剔除重复
    two_dimen_dict = {}   # 一个二维字典，用于存储每个个体所对应的hervk位点信息，后续查找个体在excel表中的位置时使用
    filelist = os.listdir(r'G:\项目-HERV\individual-analysis\merge_vcf_file')
    f1 = open('G:/项目-HERV/individual-analysis/simple_position_list.vcf', "a")
    header = 'position\tlenth\tALT\tstrand'
    f1.write(header + '\n')
    for file in filelist:
        name = file.replace("singleHERVKDISCOVERY_HERVK.final_comp.vcf", '')
        two_dimen_dict[name] = {}
        ftemp = open('G:/项目-HERV/individual-analysis/merge_vcf_file/'+file, "r")
        for i in ftemp.readlines():
            i = i.strip()
            if i.startswith("##"):
                pass
            else:
                colum = i.split("\t")
                if 'ac0' in colum[6]:
                    pass
                else:
                    if colum[0] != '#CHROM':
                        position = colum[0] + "_" + colum[1] + "-" + str(int(colum[1])+int(colum[7].split(';')[4].split('=')[1]))
                        lenth = colum[7].split(';')[4].split('=')[1]
                        ALT = colum[3]
                        strand = colum[7].split(";")[5].split(',')[3]
                        genotype = colum[9]
                        info = position + '\t' + lenth + "\t" + ALT + "\t" + strand
                        two_dimen_dict[name][info] = genotype
                        if info not in hervklist:
                            hervklist.append(info)
                            f1.write(info+'\n')
        ftemp.close()
    f1.close()
    return two_dimen_dict


def find_individual_colum(mydict):
    personlocationdict = {}
    for i in mydict:
        print(i)
        i = str(i)
        value, row2, colum = search_value(i)
        personlocationdict[i] = colum
    print(personlocationdict)


def write_0_or_1(mydict, personlocationdict):
    f1 = open('G:/项目-HERV/individual-analysis/position_list2.vcf', "r")
    file = load_workbook("G:/项目-HERV/individual-analysis/position_list2.xlsx")
    excelfile = file['Sheet1']
    row1 = 1
    for line in f1.readlines():
        print(row1)
        if line.startswith("position"):
            pass
        else:
            for i in mydict:
                i = str(i)
                colum = personlocationdict[i]
                if line.strip() in mydict[i].keys():
                    excelfile.cell(row1, colum, '1')
                else:
                    excelfile.cell(row1, colum, '0')
            file.save("G:/项目-HERV/individual-analysis/position_list2.xlsx")
        row1 += 1

def detailed_position_list():
    filelist = os.listdir(r'G:\项目-HERV\individual-analysis\merge_vcf_file')
    header = 'position\tlenth\tALT\tstrand\tfilter\tTSD\tlocation\tINTERNAL\tvariation\t'
    f1 = open('G:/项目-HERV/individual-analysis/position_list2.vcf', "a")
    f1.write(header+'\n')

    positionlist = []
    mydict = {}
    for file in filelist:
        name = file.replace("singleHERVKDISCOVERY_HERVK.final_comp.vcf", '')
        mydict[name] = {}
        ftemp = open('G:/项目-HERV/individual-analysis/merge_vcf_file/'+file, "r")
        for i in ftemp.readlines():
            i = i.strip()
            if i.startswith("##"):
                pass
            else:
                colum = i.split("\t")
                if 'ac0' in colum[6]:
                    pass
                else:
                    if colum[0] != '#CHROM':
                        position = colum[0] + "_" + colum[1] + "-" + str(int(colum[1])+int(colum[7].split(';')[4].split('=')[1]))
                        lenth = colum[7].split(';')[4].split('=')[1]
                        ALT = colum[3]
                        strand = colum[7].split(";")[5].split(',')[3]
                        filter = colum[6]
                        TSD = colum[7].split(";")[0].split('=')[1]
                        location = colum[7].split(";")[5].split(',')[1] + '--' + colum[7].split(";")[5].split(',')[2]
                        INTERNAL = colum[7].split(";")[2]
                        variation = colum[7].split(";")[6]
                        genotype = colum[9]
                        info = position+'\t'+lenth+"\t"+ALT+"\t"+strand+"\t"+filter+"\t"+TSD+"\t"+location+"\t"+INTERNAL+"\t"+variation
                        mydict[name][info] = genotype
                        if info not in positionlist:
                            positionlist.append(info)
                            f1.write(info+'\n')
        ftemp.close()
    f1.close()
    return mydict

#
# # personlocationdict = {}
# # for i in mydict:
# #     print(i)
# #     i = str(i)
# #     value, row2, colum = search_value(i)
# #     personlocationdict[i] = colum
# # print(personlocationdict)
#
# personlocationdict ={'101-5': 9, '102-5': 10, '103-5': 11, '104-5': 12, '105-5': 13, '106-5': 14, '107-5': 15, '108-5': 16, '109-5': 17, '1097': 18, '10A-9': 19, '110-5': 20, '1100': 21, '111-9': 22, '111': 23, '112': 24, '113': 25, '114': 26, '115-9': 27, '115': 28, '116-9': 29, '1166': 30, '116': 31, '117-9': 32, '117': 33, '118-9': 34, '1189': 35, '118': 36, '119-9': 37, '1192': 38, '11A-9': 39, '120-9': 40, '121-9': 41, '1211': 42, '122-9': 43, '12A-9': 44, '1344': 45, '1357': 46, '1383': 47, '13A-9': 48, '1489': 49, '14A-9': 50, '1622': 51, '1649': 52, '1651': 53, '16A-9': 54, '1726': 55, '1761': 56, '17A-9': 57, '1806': 58, '1807': 59, '1822': 60, '1825': 61, '1831': 62, '1835': 63, '1843': 64, '1851': 65, '1859': 66, '1860': 67, '1867': 68, '1876': 69, '1885': 70, '1886': 71, '1896': 72, '18A-9': 73, '1926': 74, '1938': 75, '1955': 76, '1958': 77, '1960': 78, '1967': 79, '19A-9': 80, '2006': 81, '2020': 82, '20A-9': 83, '21A-9': 84, '2A-9': 85, '300-9': 86, '301-9': 87, '302-9': 88, '303-9': 89, '304-9': 90, '305-9': 91, '306-9': 92, '307-9': 93, '308-9': 94, '309-9': 95, '310-9': 96, '312-9': 97, '313-9': 98, '314-9': 99, '315-9': 100, '316-9': 101, '317-9': 102, '318-9': 103, '319-9': 104, '367': 105, '379': 106, '383': 107, '385': 108, '394': 109, '395': 110, '397': 111, '3A-9': 112, '402': 113, '403': 114, '404': 115, '405': 116, '406': 117, '408': 118, '415': 119, '421': 120, '422': 121, '424': 122, '425': 123, '428': 124, '433': 125, '439': 126, '444': 127, '455': 128, '4A-9': 129, '511': 130, '573': 131, '590': 132, '5A-9': 133, '632': 134, '6A-9': 135, '7A-9': 136, '8A-9': 137, '920': 138, '9A-9': 139}
#
#
# file = load_workbook("G:/项目-HERV/individual-analysis/position_list.xlsx")
# excelfile = file['Sheet1']
# row1 = 1
# for line in f1.readlines():
#     print(row1)
#     if line.startswith("position"):
#         pass
#     else:
#         for i in mydict:
#             i = str(i)
#             colum = personlocationdict[i]
#             if line.strip() in mydict[i].keys():
#                 excelfile.cell(row1, colum, '1')
#             else:
#                 excelfile.cell(row1, colum, '0')
#         file.save("G:/项目-HERV/individual-analysis/position_list.xlsx")
#
#     row1 += 1


# 主程序
if __name__ == "__main__":
    # two_dimen_dict = simple_position_list()
    # # find_individual_colum(two_dimen_dict)
    # person_colum_dict = {'101-5': 5, '102-5': 6, '103-5': 7, '104-5': 8, '105-5': 9, '106-5': 10, '107-5': 11, '108-5': 12, '109-5': 13,
    #  '1097': 14, '10A-9': 15, '110-5': 16, '1100': 17, '111-9': 18, '111': 19, '112': 20, '113': 21, '114': 22,
    #  '115-9': 23, '115': 24, '116-9': 25, '1166': 26, '116': 27, '117-9': 28, '117': 29, '118-9': 30, '1189': 31,
    #  '118': 32, '119-9': 33, '1192': 34, '11A-9': 35, '120-9': 36, '121-9': 37, '1211': 38, '122-9': 39, '12A-9': 40,
    #  '1344': 41, '1357': 42, '1383': 43, '13A-9': 44, '1489': 45, '14A-9': 46, '1622': 47, '1649': 48, '1651': 49,
    #  '16A-9': 50, '1726': 51, '1761': 52, '17A-9': 53, '1806': 54, '1807': 55, '1822': 56, '1825': 57, '1831': 58,
    #  '1835': 59, '1843': 60, '1851': 61, '1859': 62, '1860': 63, '1867': 64, '1876': 65, '1885': 66, '1886': 67,
    #  '1896': 68, '18A-9': 69, '1926': 70, '1938': 71, '1955': 72, '1958': 73, '1960': 74, '1967': 75, '19A-9': 76,
    #  '2006': 77, '2020': 78, '20A-9': 79, '21A-9': 80, '2A-9': 81, '300-9': 82, '301-9': 83, '302-9': 84, '303-9': 85,
    #  '304-9': 86, '305-9': 87, '306-9': 88, '307-9': 89, '308-9': 90, '309-9': 91, '310-9': 92, '312-9': 93,
    #  '313-9': 94, '314-9': 95, '315-9': 96, '316-9': 97, '317-9': 98, '318-9': 99, '319-9': 100, '367': 101, '379': 102,
    #  '383': 103, '385': 104, '394': 105, '395': 106, '397': 107, '3A-9': 108, '402': 109, '403': 110, '404': 111,
    #  '405': 112, '406': 113, '408': 114, '415': 115, '421': 116, '422': 117, '424': 118, '425': 119, '428': 120,
    #  '433': 121, '439': 122, '444': 123, '455': 124, '4A-9': 125, '511': 126, '573': 127, '590': 128, '5A-9': 129,
    #  '632': 130, '6A-9': 131, '7A-9': 132, '8A-9': 133, '920': 134, '9A-9': 135}
    # write_0_or_1(two_dimen_dict,person_colum_dict)

    mydict = detailed_position_list()
    # find_individual_colum(mydict)
    person_colum_dict = {'101-5': 10, '102-5': 11, '103-5': 12, '104-5': 13, '105-5': 14, '106-5': 15, '107-5': 16, '108-5': 17,
     '109-5': 18, '1097': 19, '10A-9': 20, '110-5': 21, '1100': 22, '111-9': 23, '111': 24, '112': 25, '113': 26,
     '114': 27, '115-9': 28, '115': 29, '116-9': 30, '1166': 31, '116': 32, '117-9': 33, '117': 34, '118-9': 35,
     '1189': 36, '118': 37, '119-9': 38, '1192': 39, '11A-9': 40, '120-9': 41, '121-9': 42, '1211': 43, '122-9': 44,
     '12A-9': 45, '1344': 46, '1357': 47, '1383': 48, '13A-9': 49, '1489': 50, '14A-9': 51, '1622': 52, '1649': 53,
     '1651': 54, '16A-9': 55, '1726': 56, '1761': 57, '17A-9': 58, '1806': 59, '1807': 60, '1822': 61, '1825': 62,
     '1831': 63, '1835': 64, '1843': 65, '1851': 66, '1859': 67, '1860': 68, '1867': 69, '1876': 70, '1885': 71,
     '1886': 72, '1896': 73, '18A-9': 74, '1926': 75, '1938': 76, '1955': 77, '1958': 78, '1960': 79, '1967': 80,
     '19A-9': 81, '2006': 82, '2020': 83, '20A-9': 84, '21A-9': 85, '2A-9': 86, '300-9': 87, '301-9': 88, '302-9': 89,
     '303-9': 90, '304-9': 91, '305-9': 92, '306-9': 93, '307-9': 94, '308-9': 95, '309-9': 96, '310-9': 97,
     '312-9': 98, '313-9': 99, '314-9': 100, '315-9': 101, '316-9': 102, '317-9': 103, '318-9': 104, '319-9': 105,
     '367': 106, '379': 107, '383': 108, '385': 109, '394': 110, '395': 111, '397': 112, '3A-9': 113, '402': 114, '403': 115,
     '404': 116, '405': 117, '406': 118, '408': 119, '415': 120, '421': 121, '422': 122, '424': 123, '425': 124, '428': 125,
     '433': 126, '439': 127, '444': 128, '455': 129, '4A-9': 130, '511': 131, '573': 132, '590': 133, '5A-9': 134, '632': 135,
     '6A-9': 136, '7A-9': 137, '8A-9': 138, '920': 139, '9A-9': 140}
    write_0_or_1(mydict, person_colum_dict)

